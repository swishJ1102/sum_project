import os

import javalang
from openpyxl.reader.excel import load_workbook

SOURCE_EXCEL = "C:/Users/Administrator/Desktop/total.xlsx"
SOURCE_SHEET = "BASE"

PATH_NAME = 'cis'
BASE_FOLDER = 'E:\\BIP\\1基線庫\\03source\\00.受領資料\\商品\\3.23.8\\' + PATH_NAME

JAVA_IMPORT = []

ROW_CONTENT = []
ROW_CONTENT_LIST = []
ROW_CONTENT_INDEX = 0


def find_java_files(folder_path_to_find, word):
    java_files_list = []
    for root, dirs, files in os.walk(folder_path_to_find):
        for file_con in files:
            if file_con.endswith(word + ".java"):
                java_files_list.append(os.path.join(root, file_con))
    return java_files_list


def read_source_excel(path, sheet):
    sw = load_workbook(f'{path}', data_only=True)

    try:
        src_sheet = sw[f'{sheet}']
    except KeyError:
        raise KeyError('シートが存在しません。')

    source_list = []
    for row in src_sheet.iter_rows():
        if row[0].value != 'cis':
            continue
        source = get_str_after_last_dot(row[2].value, "\\")
        # print(str(row[0].row - 1) + "|" + source)
        source_list.append(source)
    return source_list


def parse_java_file(java_code):
    try:
        tree = javalang.parse.parse(java_code)
    except Exception as e:
        print('error ::::: ', e)
        return
    parse_imports(tree)


def parse_imports(tree):
    global JAVA_IMPORT
    JAVA_IMPORT = []
    import_statements = [imp.path for imp in tree.imports]
    JAVA_IMPORT = import_statements


def count_java_code_lines(file_path):
    code_lines = 0
    in_block_comment = False
    with open(file_path, 'r', encoding='utf-8') as file:
        for line in file:
            stripped_line = line.strip()
            if in_block_comment:
                if '*/' in stripped_line:
                    in_block_comment = False
                    stripped_line = stripped_line.split('*/', 1)[1]
                else:
                    continue
            if stripped_line == '' or stripped_line.startswith('//'):
                continue
            if '/*' in stripped_line:
                in_block_comment = True
                stripped_line = stripped_line.split('/*', 1)[0]
                if '*/' in stripped_line:
                    in_block_comment = False
                    stripped_line = stripped_line.split('*/', 1)[1]
            if stripped_line:
                code_lines += 1
    return code_lines


def get_str_after_last_dot(string, split):
    return string.split(split)[len(string.split(split)) - 1]


def find_sub_source_from_import():
    global ROW_CONTENT_INDEX
    import_list = []
    for import_content in JAVA_IMPORT:
        if import_content.find("enability") < 0:
            continue
        if import_content.find("AccessLogRegisterBussiness") > 0:
            continue
        if import_content.find("common") > 0 \
                or import_content.find("Common") > 0 \
                or import_content.find('model.') > 0 \
                or import_content.find('constants.') > 0 \
                or import_content.find('entity.') > 0 \
                or import_content.find('dao.') > 0:
            continue
        import_list.append(import_content)
    print("import : ", import_list)
    if len(import_list) >= 1:
        ROW_CONTENT_INDEX = len(ROW_CONTENT) - 1
    for import_l in import_list:
        sub_java_file = find_java_files(BASE_FOLDER, get_str_after_last_dot(import_l, ".") + "Impl")
        if len(sub_java_file) == 0:
            sub_java_file = find_java_files(BASE_FOLDER, get_str_after_last_dot(import_l, "."))
            sub_java_file_path = sub_java_file[0]
            # row_content = list(ROW_CONTENT[len(ROW_CONTENT) - 1])
            row_content = list(ROW_CONTENT[ROW_CONTENT_INDEX])
            row_content.append(get_str_after_last_dot(import_l, ".") + ".java")
            row_content.append(count_java_code_lines(sub_java_file_path) / 1000)
            ROW_CONTENT.append(row_content)
            with open(sub_java_file_path, 'r', encoding="utf-8") as sub_java_file_for_parse:
                sub_java_code = sub_java_file_for_parse.read()
                parse_java_file(sub_java_code)
                find_sub_source_from_import()
        else:
            sub_java_file_path = sub_java_file[0]
            row_content = list(ROW_CONTENT[len(ROW_CONTENT) - 1])
            sub_java_file_name = get_str_after_last_dot(sub_java_file_path, "\\")
            row_content.append(sub_java_file_name.split(".")[0] + ".java")
            row_content.append(count_java_code_lines(sub_java_file_path) / 1000)
            ROW_CONTENT.append(row_content)
            with open(sub_java_file_path, 'r', encoding="utf-8") as sub_java_file_for_parse:
                sub_java_code = sub_java_file_for_parse.read()
                parse_java_file(sub_java_code)
                find_sub_source_from_import()
        # row_content = list(ROW_CONTENT[len(ROW_CONTENT) - 1])
        # row_content.append(get_str_after_last_dot(import_l, "."))
        # ROW_CONTENT.append(row_content)


java_file_list = read_source_excel(SOURCE_EXCEL, SOURCE_SHEET)

for java_file in java_file_list:

    java_files = find_java_files(BASE_FOLDER, java_file.split(".")[0])
    if len(java_files) == 0:
        pass
    else:
        java_file_path = java_files[0]
        ROW_CONTENT.append([java_file, count_java_code_lines(java_file_path) / 1000])
        print("java_file : ", get_str_after_last_dot(java_file_path, "\\"))
        with open(java_file_path, 'r', encoding="utf-8") as java_file_for_parse:
            java_code = java_file_for_parse.read()
            parse_java_file(java_code)
            find_sub_source_from_import()

for row in ROW_CONTENT:
    print(row)
